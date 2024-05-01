import re
import pandas as pd
import socket
from openpyxl import load_workbook


def extract_blocks(file_path):
    """
    从给定的txt文件中提取并结构化内容。

    参数:
    - file_path: txt文件的路径。

    返回:
    - 一个字典，键为'edit'行的内容，值为对应'edit'和'next'之间的内容列表。
    """
    structured_data = {}
    current_key = None
    with open(file_path, "r", encoding="utf-8") as file:
        for line in file:
            line = line.strip()
            if line.startswith("edit"):
                current_key = line  # 'edit'行成为新的键
                structured_data[current_key] = []
            elif line == "next":
                current_key = None  # 'next'行意味着当前块结束
            elif current_key:
                structured_data[current_key].append(line)
    return structured_data


def filter_data(structured_data):
    """
    根据结构化的内容筛选数据。

    参数:
    - structured_data: 函数1的输出结果。

    返回:
    - 一个字典列表，每个字典包含指定的7项内容。
    """
    filtered_data = []
    for edit, lines in structured_data.items():
        # 初始化存储结构
        entry = {
            "edit": extract_edit_content(edit),
            "input-device": "",
            "output-device": "",
            "src": [],
            "dst": [],
            "gateway": "",
            "comments": "",
        }

        for line in lines:
            if line.startswith("set input-device"):
                entry["input-device"] = extract_quoted_content(line)
            elif line.startswith("set output-device"):
                entry["output-device"] = extract_quoted_content(line)
            elif line.startswith("set src ") or line.startswith("set srcaddr "):
                entry["src"].extend(extract_quoted_contents(line))
            elif line.startswith("set dst ") or line.startswith("set dstaddr "):
                entry["dst"].extend(extract_quoted_contents(line))
            elif line.startswith("set gateway"):
                entry["gateway"] = line.split(" ", 2)[2]
            elif line.startswith("set comments"):
                entry["comments"] = extract_quoted_content(line)

        filtered_data.append(entry)
    return filtered_data


def extract_edit_content(edit_line):
    """
    从edit行提取内容。
    """
    match = re.search(r'edit\s*(?:"([^"]+)"|(\S+))', edit_line)
    if match:
        return match.group(1) if match.group(1) else match.group(2)


def extract_quoted_content(line):
    """
    从一行中提取双引号内的内容。
    """
    match = re.search(r'"([^"]+)"', line)
    return match.group(1) if match else ""


def extract_quoted_contents(line):
    """
    提取一行中所有双引号内的内容，支持多个。
    """
    return re.findall(r'"([^"]+)"', line)


def create_excel(filtered_data, excel_path):
    """
    根据筛选好的数据创建Excel文件并保存。

    参数:
    - filtered_data: 函数2的输出结果。
    - excel_path: Excel文件保存的路径。
    """
    # 创建一个Pandas DataFrame来存储数据
    data_rows = []
    for entry in filtered_data:
        if entry["edit"].isdigit():  # 筛选纯数字的edit内容
            row = [
                entry["edit"],
                entry["input-device"],
                entry["output-device"],
                ", ".join(entry["src"]),
                ", ".join(entry["dst"]),
                entry["gateway"],
                entry["comments"],
            ]
            data_rows.append(row)

    # 定义DataFrame列名
    df = pd.DataFrame(
        data_rows,
        columns=[
            "Item",
            "Source Interface",
            "Destination Interface",
            "Source IP",
            "Destination IP",
            "Gateway",
            "Comments",
        ],
    )

    # 使用ExcelWriter保存DataFrame到Excel文件
    with pd.ExcelWriter(excel_path, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name="Filtered Data", index=False)
        workbook = writer.book
        worksheet = writer.sheets["Filtered Data"]
        # 设置列宽
        worksheet.set_column("A:B", 20)
        worksheet.set_column("C:C", 30)
        worksheet.set_column("D:E", 40)
        worksheet.set_column("F:F", 20)
        worksheet.set_column("G:G", 40)


def filter_and_process(src_data):
    """
    根据srcaddr和dstaddr进一步筛选和处理数据。

    参数:
    - src_data: 函数1的输出结果，即原始的结构化数据。

    返回:
    - 一个字典，键为edit块内容，值为进一步处理的srcaddr和dstaddr数据。
    """
    result = {}
    for edit, lines in src_data.items():
        srcaddr_dict = {}
        dstaddr_dict = {}
        for line in lines:
            if "set srcaddr" in line or "set dstaddr" in line:
                addresses = extract_quoted_contents(line)
                if "srcaddr" in line:
                    for addr in addresses:
                        srcaddr_dict[addr] = retrieve_group_or_unit(src_data, addr)
                if "dstaddr" in line:
                    for addr in addresses:
                        dstaddr_dict[addr] = retrieve_group_or_unit(src_data, addr)
        if srcaddr_dict or dstaddr_dict:
            result[edit] = {"srcaddr": srcaddr_dict, "dstaddr": dstaddr_dict}
    return result

def nslookup(fqdn):
    """
    使用nslookup查询fqdn的IP地址。

    参数:
    - fqdn: 完全限定域名(FQDN)

    返回:
    - 查询到的IP地址列表，如果查询失败返回一个空列表。
    """
    try:
        return socket.gethostbyname_ex(fqdn)[2]
    except socket.gaierror:
        return False


def retrieve_group_or_unit(src_data, addr):
    """
    检索指定地址对应的群组或单元数据。

    参数:
    - src_data: 原始的结构化数据。
    - addr: 要检索的地址名称。

    返回:
    - 一个字典，包含地址的群组成员或单元信息。
    """
    for edit, lines in src_data.items():
        if f'edit "{addr}"' == edit.strip():
            member_data = {}
            start_ip, end_ip = None, None
            for line in lines:
                if "set member" in line:
                    members = extract_quoted_contents(line)
                    for member in members:
                        member_data[member] = retrieve_group_or_unit(src_data, member)
                elif "set subnet" in line:
                    subnet = line.split(" ", 2)[2]
                elif "set fqdn" in line:
                    fqdn = extract_quoted_content(line)
                    fqdn_ips = nslookup(fqdn)
                    fqdn_result = fqdn
                    if fqdn_ips:
                        fqdn_result = f"{fqdn}: {', '.join(fqdn_ips)}"
                elif "set start-ip" in line:
                    start_ip = line.split(" ", 2)[2]
                elif "set end-ip" in line:
                    end_ip = line.split(" ", 2)[2]
            if start_ip and end_ip:
                return f"start_ip:{start_ip}, end_ip:{end_ip}"
            elif "subnet" in locals():
                return subnet
            elif "fqdn_result" in locals():
                return fqdn_result
            if member_data:
                return member_data
    return {}


def format_detail(detail):
    """
    将详情数据格式化为字符串。
    如果数据是字典，则将其转换为“键: 值”格式的字符串。
    否则，直接转换为字符串。
    """
    if isinstance(detail, dict):
        return "; ".join([f"{k}: {v}" for k, v in detail.items()])
    return str(detail)


def dict_merge(data):
    if isinstance(data, dict):
        return "; ".join([f"{member}: {format_detail(member_data)}"
                for member, member_data in data.items()])
    return data


def add_data_to_excel(processed_data, excel_path):
    """
    根据处理好的数据往已有的Excel文件添加内容。

    参数:
    - processed_data: 函数4的输出结果。
    - excel_path: 已存在的Excel文件路径。
    """
    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a") as writer:
        for edit, data in processed_data.items():
            for addr_type in ["srcaddr", "dstaddr"]:
                rows = []
                for addr, details in data[addr_type].items():
                    if isinstance(details, dict):  # 如果是群组
                        rows.append(["", ""])
                        rows.append([addr, ""])
                        for member, member_data in details.items():
                            member_data = dict_merge(member_data)
                            rows.append([member, member_data])
                        rows.append(["", ""])
                    else:  # 如果直接是subnet或fqdn
                        rows.append([addr, details])
                if rows:  # 如果有数据，则创建sheet
                    df = pd.DataFrame(rows, columns=["Name", "Subnet/FQDN"])
                    df.to_excel(writer, sheet_name=f"{edit}_{addr_type}", index=False)

    workbook = load_workbook(excel_path)
    for edit, data in processed_data.items():
        for addr_type in ["srcaddr", "dstaddr"]:
            sheet_name = f"{edit}_{addr_type}"
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                # 设置A和B列的宽度
                worksheet.column_dimensions["A"].width = 40
                worksheet.column_dimensions["B"].width = 40
    workbook.save(excel_path)


def run(file_path, excel_path):
    """
    执行整个处理流程。

    参数:
    - file_path: 原始txt文件的路径。
    - excel_path: Excel文件的保存路径。
    """
    structured_data = extract_blocks(file_path)
    filtered_data = filter_data(structured_data)
    create_excel(filtered_data, excel_path)  # 创建Excel文件并存入筛选后的数据
    processed_data = filter_and_process(structured_data)
    add_data_to_excel(processed_data, excel_path)  # 向Excel文件添加额外的数据


# 调用run函数，传入原始文件路径和Excel文件保存路径
# run('path_to_your_txt_file.txt', 'output_data.xlsx')

if __name__ == "__main__":
    run(
        r"D:\Downloads\仁愛PolicyRouteConfig-----.txt", r"D:\Downloads\output_data.xlsx"
    )
